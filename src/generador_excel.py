"""
generador_excel.py

Genera el archivo Excel de salida a partir de los encabezados y detalles
extraídos de uno o varios PDF.

Ahora incluye una columna adicional llamada "Observaciones",
la cual contiene el bloque de observaciones del documento y se
repite en todas las filas pertenecientes a esa factura.
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

COLUMNAS = [
    ("tipo_documento", "Tipo Documento"),
    ("ncf", "NCF"),
    ("ncf_modificado", "NCF Modificado"),
    ("fecha_documento", "Fecha Documento"),
    ("numero_documento", "No. Documento"),
    ("indicador_bien_servicio", "Indicador Bien/Servicio"),
    ("fecha_vencimiento", "Fecha Vcto."),
    ("rnc_cliente", "RNC Cliente"),
    ("cliente", "Cliente"),
    ("monto", "Monto"),
    ("id_documento_detalle", "ID Documento (Nombre de archivo Origen)"),
    ("observaciones", "Observaciones"),
]

FORMATO_FECHA = "DD/MM/YYYY"
FORMATO_MONTO = "#,##0.00"

RELLENO_ENCABEZADO_COLUMNA = PatternFill(
    "solid",
    start_color="1F4E78",
    end_color="1F4E78"
)

FUENTE_ENCABEZADO_COLUMNA = Font(
    bold=True,
    color="FFFFFF"
)

RELLENO_FILA_DOCUMENTO = PatternFill(
    "solid",
    start_color="D9E1F2",
    end_color="D9E1F2"
)

FUENTE_FILA_DOCUMENTO = Font(bold=True)


def obtener_documentos_existentes(ruta_excel: str) -> set:
    """Si el Excel de salida ya existe, devuelve el conjunto de números
    de documento (columna 'ID Documento') que ya están registrados en
    él, para poder detectar duplicados antes de agregar nuevas filas.
    Si el archivo no existe todavía, devuelve un conjunto vacío."""
    ruta = Path(ruta_excel)
    if not ruta.exists():
        return set()

    wb = load_workbook(ruta)
    if "Facturacion" not in wb.sheetnames:
        return set()

    hoja = wb["Facturacion"]
    indice_id = [c for c, _ in COLUMNAS].index("id_documento_detalle") + 1

    existentes = set()
    for fila in hoja.iter_rows(min_row=2, values_only=False):
        valor = fila[indice_id - 1].value
        if valor is not None:
            existentes.add(str(valor))

    return existentes


def _construir_filas(documentos: list) -> list:

    filas = []

    for documento in documentos:

        encabezado = documento["encabezado"]
        detalle = documento["detalle"]

        # -----------------------------
        # Fila de encabezado
        # -----------------------------

        fila_encabezado = {
            clave: encabezado.get(clave)
            for clave, _ in COLUMNAS
        }

        fila_encabezado["monto"] = encabezado.get("total")
        fila_encabezado["id_documento_detalle"] = encabezado.get(
            "numero_documento"
        )

        fila_encabezado["observaciones"] = encabezado.get(
            "observaciones"
        )

        fila_encabezado["_es_encabezado"] = True

        filas.append(fila_encabezado)

        # -----------------------------
        # Filas de detalle
        # -----------------------------

        for linea in detalle:

            fila_detalle = {
                clave: None
                for clave, _ in COLUMNAS
            }

            fila_detalle["ncf"] = linea.get("descripcion")

            fila_detalle["monto"] = linea.get("monto")

            fila_detalle["id_documento_detalle"] = linea.get(
                "numero_documento"
            )

            # Repetir observaciones


            fila_detalle["_es_encabezado"] = False

            filas.append(fila_detalle)

    return filas


def _escribir_hoja(wb: Workbook, filas: list, hoja_existente: bool = False):

    hoja = wb.active

    hoja.title = "Facturacion"

    if not hoja_existente:

        encabezados = [titulo for _, titulo in COLUMNAS]

        hoja.append(encabezados)

        for celda in hoja[1]:

            celda.font = FUENTE_ENCABEZADO_COLUMNA
            celda.fill = RELLENO_ENCABEZADO_COLUMNA
            celda.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

    indice_fecha_doc = [
        c for c, _ in COLUMNAS
    ].index("fecha_documento") + 1

    indice_fecha_vto = [
        c for c, _ in COLUMNAS
    ].index("fecha_vencimiento") + 1

    indice_monto = [
        c for c, _ in COLUMNAS
    ].index("monto") + 1

    indice_id_documento = [
        c for c, _ in COLUMNAS
    ].index("id_documento_detalle") + 1

    for fila in filas:

        valores = [
            fila.get(clave)
            for clave, _ in COLUMNAS
        ]

        hoja.append(valores)

        num_fila = hoja.max_row

        hoja.cell(
            row=num_fila,
            column=indice_fecha_doc
        ).number_format = FORMATO_FECHA

        hoja.cell(
            row=num_fila,
            column=indice_fecha_vto
        ).number_format = FORMATO_FECHA

        hoja.cell(
            row=num_fila,
            column=indice_monto
        ).number_format = FORMATO_MONTO

        # Centrar únicamente la columna ID Documento
        hoja.cell(
            row=num_fila,
            column=indice_id_documento
        ).alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        if fila.get("_es_encabezado"):

            for columna in range(
                1,
                len(COLUMNAS) + 1
            ):

                celda = hoja.cell(
                    row=num_fila,
                    column=columna
                )

                celda.fill = RELLENO_FILA_DOCUMENTO
                celda.font = FUENTE_FILA_DOCUMENTO

    for indice, (_, titulo) in enumerate(
        COLUMNAS,
        start=1
    ):

        letra = get_column_letter(indice)

        if titulo == "NCF":

            ancho = 40

        elif titulo == "Observaciones":

            ancho = 80

        elif titulo == "ID Documento (Nombre de archivo Origen)":

            ancho = 30

        else:

            ancho = max(
                len(titulo) + 4,
                14
            )

        hoja.column_dimensions[letra].width = ancho

    hoja.freeze_panes = "A2"


def generar_excel(documentos: list, ruta_salida: str):
    """Genera o ACTUALIZA el Excel de salida.

    Si `ruta_salida` ya existe, se carga el archivo y las filas nuevas
    se agregan debajo de las que ya había (acumulativo entre
    ejecuciones). Si no existe, se crea desde cero como antes.

    Esta función asume que `documentos` ya viene filtrado (sin
    duplicados respecto al Excel existente); ese filtrado lo hace
    `procesar.py` antes de llamar a esta función."""
    ruta = Path(ruta_salida)
    hoja_existente = ruta.exists()

    if hoja_existente:
        wb = load_workbook(ruta)
    else:
        wb = Workbook()

    filas = _construir_filas(documentos)

    _escribir_hoja(
        wb,
        filas,
        hoja_existente=hoja_existente,
    )

    wb.save(ruta_salida)


if __name__ == "__main__":

    import pdfplumber

    from extractor_encabezado import extraer_encabezado
    from extractor_detalle import extraer_detalle

    with pdfplumber.open("../muestras/800001168.pdf") as pdf:

        page = pdf.pages[0]

        texto = page.extract_text()

        palabras = page.extract_words()

    encabezado = extraer_encabezado(texto)

    detalle = extraer_detalle(
        palabras,
        numero_documento=encabezado["numero_documento"]
    )

    documentos = [
        {
            "encabezado": encabezado,
            "detalle": detalle
        }
    ]

    generar_excel(
        documentos,
        "../salida/resultado.xlsx"
    )

    print("Excel generado correctamente.")